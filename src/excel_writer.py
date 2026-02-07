"""
Excel report generation module using openpyxl - FIXED VERSION
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
import numpy as np
import logging
import os

class ExcelReportWriter:
    """Creates comprehensive Excel reports with dashboards and charts."""
    
    def __init__(self, config, logger=None):
        self.config = config
        self.logger = logger or logging.getLogger(__name__)
        self.wb = None
        self.styles = {}
        self._initialize_styles()
    
    def _initialize_styles(self):
        """Initialize reusable cell styles."""
        # Colors from config
        colors = self.config.COLORS
        
        # Header style
        self.styles['header'] = {
            'fill': PatternFill(start_color=colors['header_blue'], end_color=colors['header_blue'], fill_type="solid"),
            'font': Font(color="FFFFFF", bold=True, size=12),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        }
        
        # Subheader style
        self.styles['subheader'] = {
            'fill': PatternFill(start_color=colors['dark_blue'], end_color=colors['dark_blue'], fill_type="solid"),
            'font': Font(color="FFFFFF", bold=True, size=11),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        }
        
        # Data style
        self.styles['data'] = {
            'font': Font(size=10),
            'alignment': Alignment(vertical="center"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        }
        
        # Currency style
        self.styles['currency'] = {
            'font': Font(size=10),
            'alignment': Alignment(horizontal="right", vertical="center"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin')),
            'number_format': '"$"#,##0.00'
        }
        
        # Alert styles
        self.styles['critical'] = {
            'fill': PatternFill(start_color=colors['urgent_red'], end_color=colors['urgent_red'], fill_type="solid"),
            'font': Font(bold=True)
        }
        
        self.styles['warning'] = {
            'fill': PatternFill(start_color=colors['warning_yellow'], end_color=colors['warning_yellow'], fill_type="solid"),
            'font': Font(bold=True)
        }
        
        self.styles['good'] = {
            'fill': PatternFill(start_color=colors['good_green'], end_color=colors['good_green'], fill_type="solid"),
            'font': Font(bold=True)
        }
    
    def create_report(self, data_dict, analyzer, output_path):
        """Create the complete Excel report."""
        self.logger.info(f"Creating Excel report: {output_path}")
        
        try:
            self.wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.wb.sheetnames:
                default_sheet = self.wb['Sheet']
                self.wb.remove(default_sheet)
            
            # Create all sheets
            self._create_executive_dashboard(analyzer)
            
            # Only create sheets if we have data
            if 'sales' in data_dict and data_dict['sales'] is not None:
                self._create_sales_analysis_sheet(data_dict['sales'], analyzer)
            
            if 'inventory' in data_dict and data_dict['inventory'] is not None:
                self._create_inventory_sheet(data_dict['inventory'], analyzer)
            
            if 'advertising' in data_dict and data_dict['advertising'] is not None:
                self._create_advertising_sheet(data_dict['advertising'], analyzer)
            
            if 'reviews' in data_dict and data_dict['reviews'] is not None:
                self._create_reviews_sheet(data_dict['reviews'], analyzer)
            
            if analyzer.alerts:
                self._create_alerts_sheet(analyzer.alerts)
            
            # Add raw data if available
            for data_type, df in data_dict.items():
                if df is not None and len(df) > 0:
                    self._add_raw_data_sheet(df, data_type)
            
            # Save the workbook
            self.wb.save(output_path)
            self.logger.info(f"Report saved successfully: {output_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating Excel report: {e}")
            raise
    
    def _create_executive_dashboard(self, analyzer):
        """Create the executive dashboard sheet."""
        self.logger.info("Creating executive dashboard...")
        
        ws = self.wb.create_sheet(title=self.config.DEFAULT_SHEET_NAMES['dashboard'])
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "📊 Amazon Seller Executive Dashboard"
        title_cell.font = Font(size=18, bold=True, color=self.config.COLORS['dark_blue'])
        title_cell.alignment = Alignment(horizontal='center')
        
        # Subtitle
        ws.merge_cells('A2:H2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        subtitle_cell.font = Font(size=10, italic=True)
        subtitle_cell.alignment = Alignment(horizontal='center')
        
        # Get KPI summary
        kpis = analyzer.get_kpi_summary()
        
        # KPI Cards (Row 4-6)
        kpi_cards = []
        
        # Add available KPIs
        if kpis.get('total_revenue', 0) > 0:
            kpi_cards.append(["💰 Total Revenue", f"${kpis.get('total_revenue', 0):,.2f}", "sales"])
        
        if kpis.get('total_transactions', 0) > 0:
            kpi_cards.append(["🛒 Total Transactions", f"{kpis.get('total_transactions', 0):,}", "sales"])
        
        if kpis.get('avg_order_value', 0) > 0:
            kpi_cards.append(["📦 Avg Order Value", f"${kpis.get('avg_order_value', 0):,.2f}", "sales"])
        
        if kpis.get('total_products', 0) > 0:
            kpi_cards.append(["📊 Total Products", f"{kpis.get('total_products', 0):,}", "inventory"])
        
        if kpis.get('low_stock_products', 0) > 0:
            kpi_cards.append(["⚠️ Low Stock Items", f"{kpis.get('low_stock_products', 0):,}", "inventory"])
        
        if kpis.get('total_reviews', 0) > 0:
            kpi_cards.append(["⭐ Total Reviews", f"{kpis.get('total_reviews', 0):,}", "reviews"])
        
        if kpis.get('avg_rating', 0) > 0:
            kpi_cards.append(["📈 Avg Rating", f"{kpis.get('avg_rating', 0):.1f}/5", "reviews"])
        
        if kpis.get('total_ad_spend', 0) > 0:
            kpi_cards.append(["📢 Ad Spend", f"${kpis.get('total_ad_spend', 0):,.2f}", "advertising"])
        
        # Layout KPI cards
        row = 4
        for i, (title, value, category) in enumerate(kpi_cards):
            col = (i % 4) * 2 + 1  # 2 columns per card, 4 cards per row
            if i >= 4:  # Start new row after 4 cards
                row = 7
                col = ((i-4) % 4) * 2 + 1
            
            # Card title
            ws.cell(row=row, column=col, value=title).font = Font(bold=True, size=11)
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            
            # Card value
            ws.cell(row=row+1, column=col, value=value).font = Font(bold=True, size=14, color=self.config.COLORS['dark_blue'])
            ws.cell(row=row+1, column=col).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            
            # Apply color based on category
            fill_color = self._get_category_color(category)
            for r in range(row, row+2):
                for c in range(col, col+2):
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
                    ws.cell(row=r, column=c).border = self.styles['data']['border']
        
        # Alerts Section (Row 10+)
        row = 10
        ws.cell(row=row, column=1, value="🚨 Critical Alerts & Actions Required").font = Font(bold=True, size=14, color="FF0000")
        row += 1
        
        if analyzer.alerts:
            for i, alert in enumerate(analyzer.alerts[:5]):  # Show top 5 alerts
                ws.cell(row=row+i, column=1, value=f"• {alert['title']}: {alert['message']}")
                if alert['level'] == 'critical':
                    ws.cell(row=row+i, column=1).font = Font(bold=True, color="FF0000")
                elif alert['level'] == 'warning':
                    ws.cell(row=row+i, column=1).font = Font(bold=True, color="FF9900")
        
        # Top Products Section (Row 17+)
        row = 17
        ws.cell(row=row, column=1, value="🏆 Top Performing Products").font = Font(bold=True, size=14, color=self.config.COLORS['dark_blue'])
        row += 1
        
        if 'sales' in analyzer.results and 'top_products' in analyzer.results['sales']:
            top_products = analyzer.results['sales']['top_products']
            if hasattr(top_products, 'head'):
                top_products = top_products.head(5)
            
            if not top_products.empty:
                # Headers
                headers = ["Product Name", "Revenue"]
                if 'total_quantity' in top_products.columns:
                    headers.append("Quantity")
                
                for j, header in enumerate(headers):
                    cell = ws.cell(row=row, column=1+j, value=header)
                    for style_key, style_value in self.styles['header'].items():
                        setattr(cell, style_key, style_value)
                
                row += 1
                
                # Data
                for i, (_, product) in enumerate(top_products.iterrows()):
                    col_idx = 0
                    product_name = str(product.get('product_name', 'Unknown'))[:50]
                    ws.cell(row=row+i, column=col_idx+1, value=product_name)
                    col_idx += 1
                    
                    revenue = product.get('total_revenue', 0)
                    ws.cell(row=row+i, column=col_idx+1, value=revenue)
                    ws.cell(row=row+i, column=col_idx+1).number_format = '"$"#,##0.00'
                    col_idx += 1
                    
                    if 'total_quantity' in top_products.columns:
                        quantity = product.get('total_quantity', 0)
                        ws.cell(row=row+i, column=col_idx+1, value=quantity)
                        col_idx += 1
                    
                    # Apply data style
                    for j in range(1, col_idx+1):
                        cell = ws.cell(row=row+i, column=j)
                        for style_key, style_value in self.styles['data'].items():
                            if style_key != 'number_format' or j != 2:
                                setattr(cell, style_key, style_value)
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_sales_analysis_sheet(self, sales_df, analyzer):
        """Create sales analysis sheet."""
        self.logger.info("Creating sales analysis sheet...")
        
        ws = self.wb.create_sheet(title=self.config.DEFAULT_SHEET_NAMES['sales'])
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'].value = "💰 Sales Performance Analysis"
        ws['A1'].font = Font(size=16, bold=True, color=self.config.COLORS['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Summary metrics
        if 'sales' in analyzer.results:
            sales_metrics = analyzer.results['sales']
            
            metrics = []
            if 'total_revenue' in sales_metrics:
                metrics.append(["Total Revenue", f"${sales_metrics.get('total_revenue', 0):,.2f}"])
            
            if 'total_transactions' in sales_metrics:
                metrics.append(["Total Transactions", f"{sales_metrics.get('total_transactions', 0):,}"])
            
            if 'total_quantity' in sales_metrics and sales_metrics['total_quantity'] > 0:
                metrics.append(["Total Quantity Sold", f"{sales_metrics.get('total_quantity', 0):,}"])
            
            if 'avg_order_value' in sales_metrics:
                metrics.append(["Average Order Value", f"${sales_metrics.get('avg_order_value', 0):,.2f}"])
            
            if 'monthly_growth' in sales_metrics:
                metrics.append(["Monthly Growth", f"{sales_metrics.get('monthly_growth', 0):.1f}%"])
            
            row = 3
            for i, (label, value) in enumerate(metrics):
                ws.cell(row=row+i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row+i, column=2, value=value)
    
    def _create_inventory_sheet(self, inventory_df, analyzer):
        """Create inventory health sheet."""
        self.logger.info("Creating inventory health sheet...")
        
        ws = self.wb.create_sheet(title=self.config.DEFAULT_SHEET_NAMES['inventory'])
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'].value = "📦 Inventory Health Analysis"
        ws['A1'].font = Font(size=16, bold=True, color=self.config.COLORS['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if 'inventory' in analyzer.results:
            inv_metrics = analyzer.results['inventory']
            
            # Summary
            row = 3
            summary = []
            
            if 'total_products' in inv_metrics:
                summary.append(["Total Products", f"{inv_metrics.get('total_products', 0):,}"])
            
            if 'total_stock' in inv_metrics:
                summary.append(["Total Stock", f"{inv_metrics.get('total_stock', 0):,}"])
            
            if 'low_stock_count' in inv_metrics:
                summary.append(["Low Stock Items", f"{inv_metrics.get('low_stock_count', 0):,}"])
            
            if 'overstock_count' in inv_metrics:
                summary.append(["Overstock Items", f"{inv_metrics.get('overstock_count', 0):,}"])
            
            for i, (label, value) in enumerate(summary):
                ws.cell(row=row+i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row+i, column=2, value=value)
    
    def _create_advertising_sheet(self, advertising_df, analyzer):
        """Create advertising ROI sheet."""
        self.logger.info("Creating advertising analysis sheet...")
        
        ws = self.wb.create_sheet(title=self.config.DEFAULT_SHEET_NAMES['advertising'])
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'].value = "📢 Advertising Performance"
        ws['A1'].font = Font(size=16, bold=True, color=self.config.COLORS['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if 'advertising' in analyzer.results:
            ad_metrics = analyzer.results['advertising']
            
            # Summary
            row = 3
            summary = []
            
            if 'total_spend' in ad_metrics:
                summary.append(["Total Spend", f"${ad_metrics.get('total_spend', 0):,.2f}"])
            
            if 'total_clicks' in ad_metrics:
                summary.append(["Total Clicks", f"{ad_metrics.get('total_clicks', 0):,}"])
            
            if 'total_impressions' in ad_metrics:
                summary.append(["Total Impressions", f"{ad_metrics.get('total_impressions', 0):,}"])
            
            if 'avg_cpc' in ad_metrics:
                summary.append(["Average CPC", f"${ad_metrics.get('avg_cpc', 0):,.2f}"])
            
            if 'ctr' in ad_metrics:
                summary.append(["CTR", f"{ad_metrics.get('ctr', 0):.2f}%"])
            
            for i, (label, value) in enumerate(summary):
                ws.cell(row=row+i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row+i, column=2, value=value)
    
    def _create_reviews_sheet(self, reviews_df, analyzer):
        """Create customer reviews sheet."""
        self.logger.info("Creating customer reviews sheet...")
        
        ws = self.wb.create_sheet(title=self.config.DEFAULT_SHEET_NAMES['reviews'])
        
        # Title
        ws.merge_cells('A1:C1')
        ws['A1'].value = "⭐ Customer Reviews Analysis"
        ws['A1'].font = Font(size=16, bold=True, color=self.config.COLORS['dark_blue'])
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if 'reviews' in analyzer.results:
            review_metrics = analyzer.results['reviews']
            
            # Summary
            row = 3
            summary = []
            
            if 'total_reviews' in review_metrics:
                summary.append(["Total Reviews", f"{review_metrics.get('total_reviews', 0):,}"])
            
            if 'avg_rating' in review_metrics:
                summary.append(["Average Rating", f"{review_metrics.get('avg_rating', 0):.1f}/5.0"])
            
            if 'low_rated_count' in review_metrics:
                summary.append(["Low Ratings (<3.0)", f"{review_metrics.get('low_rated_count', 0):,}"])
            
            for i, (label, value) in enumerate(summary):
                ws.cell(row=row+i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row+i, column=2, value=value)
    
    def _create_alerts_sheet(self, alerts):
        """Create alerts sheet."""
        self.logger.info("Creating alerts sheet...")
        
        ws = self.wb.create_sheet(title=self.config.DEFAULT_SHEET_NAMES['alerts'])
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'].value = "🚨 Actionable Alerts"
        ws['A1'].font = Font(size=16, bold=True, color="FF0000")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Priority", "Category", "Title", "Message"]
        for j, header in enumerate(headers):
            cell = ws.cell(row=3, column=1+j, value=header)
            for style_key, style_value in self.styles['header'].items():
                setattr(cell, style_key, style_value)
        
        # Data
        for i, alert in enumerate(alerts):
            row = 4 + i
            
            # Priority with color coding
            priority_cell = ws.cell(row=row, column=1, value=alert['level'].upper())
            if alert['level'] == 'critical':
                priority_cell.fill = self.styles['critical']['fill']
                priority_cell.font = Font(bold=True, color="FFFFFF")
            elif alert['level'] == 'warning':
                priority_cell.fill = self.styles['warning']['fill']
            
            # Other columns
            ws.cell(row=row, column=2, value=alert['type'].upper())
            ws.cell(row=row, column=3, value=alert['title'])
            ws.cell(row=row, column=4, value=alert['message'])
            
            # Apply data style
            for j in range(1, 5):
                cell = ws.cell(row=row, column=j)
                for style_key, style_value in self.styles['data'].items():
                    if style_key not in ['fill', 'font'] or j != 1:
                        setattr(cell, style_key, style_value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 60
    
    def _add_raw_data_sheet(self, df, data_type):
        """Add raw data sheet."""
        self.logger.info(f"Adding raw data sheet for {data_type}...")
        
        # Clean sheet name
        sheet_name = f"Raw {data_type.title()}"
        if len(sheet_name) > 31:  # Excel sheet name limit
            sheet_name = sheet_name[:31]
        
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Add headers
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=1, column=j+1, value=col)
            for style_key, style_value in self.styles['header'].items():
                setattr(cell, style_key, style_value)
        
        # Add data
        for i, row in df.iterrows():
            for j, col in enumerate(df.columns):
                value = row[col]
                cell = ws.cell(row=i+2, column=j+1, value=value)
                
                # Apply data style
                for style_key, style_value in self.styles['data'].items():
                    if style_key != 'number_format':
                        setattr(cell, style_key, style_value)
        
        # Auto-filter
        max_row = len(df) + 1
        max_col = len(df.columns)
        if max_row > 1 and max_col > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        
        # Adjust column widths
        for j, col in enumerate(df.columns):
            col_letter = get_column_letter(j+1)
            max_length = 0
            try:
                max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2
            except:
                max_length = len(col) + 2
            ws.column_dimensions[col_letter].width = min(max_length, 50)
    
    def _get_category_color(self, category):
        """Get color for category."""
        color_map = {
            'sales': 'BDD7EE',  # Light blue
            'inventory': 'C6EFCE',  # Light green
            'reviews': 'FFEB9C',  # Light yellow
            'advertising': 'F8CBAD'  # Light orange
        }
        return color_map.get(category, 'E7E6E6')  # Default gray
    
    def _format_column_name(self, col_name):
        """Format column name for display."""
        return str(col_name).replace('_', ' ').title()
